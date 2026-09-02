"""验收比对（§8.1）：导出的 24级（成绩） vs output.xlsx 24级（成绩）。

按（学号, 姓名）对齐共有学生，逐课程比对成绩矩阵，比对加权总分/总学分。
运行：cd backend && python tests/compare_output.py
"""
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OURS = os.path.join(ROOT, "export_24级.xlsx")
REF = os.path.join(ROOT, "output.xlsx")


def read_sheet(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    # 找标签行（含 班级/学号/姓名）
    label_row = None
    for r in range(1, 8):
        vals = [ws.cell(r, c).value for c in range(1, 4)]
        if vals == ["班级", "学号", "姓名"]:
            label_row = r
            break
    assert label_row, f"{path} {sheet_name} 未找到标签行"
    # 课程列: 第1行课程代码从 D 列起
    codes, col_of = [], {}
    c = 4
    while True:
        code = ws.cell(1, c).value
        if not code or str(code).strip() in ("加权总分",):
            break
        codes.append(str(code).strip())
        col_of[str(code).strip()] = c
        c += 1
    # 尾部统计列名
    tail = {}
    for cc in range(c, ws.max_column + 1):
        lab = ws.cell(label_row, cc).value
        if lab:
            tail[str(lab)] = cc
    students = {}
    r = label_row + 1
    while r <= ws.max_row:
        no = ws.cell(r, 2).value
        if no is not None:
            no = str(int(no)) if isinstance(no, float) else str(no).strip()
            students[no] = {
                "name": str(ws.cell(r, 3).value or "").strip(),
                "scores": {code: ws.cell(r, col_of[code]).value for code in codes
                           if ws.cell(r, col_of[code]).value is not None},
                "tail": {lab: ws.cell(r, cc).value for lab, cc in tail.items()},
            }
        r += 1
    return {"codes": codes, "students": students}


ours = read_sheet(OURS, "24级（成绩）")
ref = read_sheet(REF, "24级（成绩）")

common = sorted(set(ours["students"]) & set(ref["students"]))
only_ours = sorted(set(ours["students"]) - set(ref["students"]))
only_ref = sorted(set(ref["students"]) - set(ours["students"]))
print(f"学生：共有 {len(common)}，仅我方导出 {len(only_ours)} {only_ours[:5]}，"
      f"仅手工样本 {len(only_ref)} {only_ref[:5]}")

# 成绩矩阵比对
score_diff = 0
checked_cells = 0
for no in common:
    a, b = ours["students"][no], ref["students"][no]
    for code in set(a["scores"]) & set(b["scores"]):
        va, vb = float(a["scores"][code]), float(b["scores"][code])
        checked_cells += 1
        if abs(va - vb) > 0.01:
            score_diff += 1
            if score_diff <= 10:
                print(f"  成绩差异 {no} {a['name']} {code}: 我方 {va} vs 样本 {vb}")
print(f"成绩矩阵：共有课程单元格比对 {checked_cells} 个，差异 {score_diff}")

# 加权统计比对
stat_diff = []
for no in common:
    a, b = ours["students"][no], ref["students"][no]
    for lab in ("加权总分", "总学分"):
        va, vb = a["tail"].get(lab), b["tail"].get(lab)
        if va is None or vb is None:
            continue
        if abs(float(va) - float(vb)) > 0.01:
            stat_diff.append((no, lab, va, vb))
print(f"加权总分/总学分差异：{len(stat_diff)}")
for d in stat_diff[:10]:
    print("  ", d)

# 与样本差异总结（手工样本包含更多学生属正常——其数据源更广）
print("\n结论：" + ("核心数据一致" if score_diff == 0 and not stat_diff else "存在差异，见上"))

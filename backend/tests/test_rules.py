"""单元测试（设计文档 §10.15 / §8）：换算、解析、软校验、排名、补考规则、导出结构。

运行：cd backend && python -m pytest tests/ -q
"""
import io
import os
import sys
from datetime import datetime
from unittest import mock

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.services.convert import convert_level, parse_number, sum_detail_terms  # noqa: E402
from app.services.calc import (DEFAULT_ITEMS, _competition_rank,  # noqa: E402
                               collapse_retakes, compute_rankings)
from app.services.calc import StudentResult  # noqa: E402
from app.services.eval_import import _squash_class_key, parse_eval_workbook  # noqa: E402
from app.services.score_import import (infer_enrollment_year,  # noqa: E402
                                       infer_grade_name, normalize_class_name,
                                       parse_score_workbook)
from app.models import ScoreRecord  # noqa: E402

CONVERSION = {"优": 95, "优秀": 95, "良": 85, "良好": 85, "中": 75, "中等": 75,
              "及格": 65, "不及格": 0, "合格": 80, "不合格": 0}


# ---------- 1. 等级换算 ----------
def test_convert_level():
    assert convert_level("优", CONVERSION) == 95
    assert convert_level("良好", CONVERSION) == 85
    # 「合格」与「及格」严格区分（§9）
    assert convert_level("合格", CONVERSION) == 80
    assert convert_level("及格", CONVERSION) == 65
    assert convert_level("缓考", CONVERSION) is None


def test_parse_number():
    assert parse_number("90") == 90.0
    assert parse_number(88.5) == 88.5
    assert parse_number("") is None
    assert parse_number("缓考") is None


# ---------- 2. 明细软校验求和 ----------
def test_sum_detail_terms():
    # 普通加减分
    assert sum_detail_terms("基础分+23\n学校社会实践重点团队+0.5") == 23.5
    # 「六级成绩587分+5」：无符号的 587 不计，+5 计入（§10.7）
    assert sum_detail_terms("六级成绩587分+5") == 5
    # 「525心理嘉年华+0.5」同理
    assert sum_detail_terms("参加525“漫步心域”心理嘉年华+0.5") == 0.5
    # 减分与年份区间
    assert sum_detail_terms("违纪-2\n参加2024-2025学年活动+1") == -1
    assert sum_detail_terms("无") == 0


# ---------- 3. 班级/年级推断 ----------
def test_class_and_grade_inference():
    assert normalize_class_name("建筑类2402班") == "建筑类2402"
    assert infer_grade_name("建筑类2401") == "24级"
    assert infer_grade_name("建筑2402班") == "24级"
    assert infer_enrollment_year("24级") == 2024
    # 班级名变体归一（§4.2）
    assert _squash_class_key("建筑类2402班") == _squash_class_key("建筑 2402")
    assert _squash_class_key("建筑类2402") == _squash_class_key("建筑2402班")


# ---------- 4. 1.xls 解析（用真实样本） ----------
SAMPLE1 = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "..", "1.xls")


def _sample1_path():
    p = os.path.abspath(SAMPLE1)
    return p if os.path.exists(p) else None


@pytest.mark.skipif(_sample1_path() is None, reason="样本 1.xls 不存在")
def test_parse_score_workbook_sample():
    with open(_sample1_path(), "rb") as f:
        parsed = parse_score_workbook("1.xls", f.read(), CONVERSION)
    assert parsed.year_names == ["2024-2025"]
    assert set(parsed.class_names) == {"建筑类2401", "建筑类2402"}
    assert parsed.student_count == 50  # 样本实际 50 名学生（两班）
    assert parsed.course_count > 20
    # 等级文本被换算为数字；数字保持
    sample_num = [r for r in parsed.rows if r.score_num == 95 and r.score_raw == "优"]
    assert sample_num
    # 未匹配等级（若有）不计入 score_num
    for r in parsed.rows:
        if r.score_raw in CONVERSION:
            assert r.score_num == CONVERSION[r.score_raw]


# ---------- 5. 补考/重修取分规则 ----------
def _rec(code, sem, num, rid=0, gpa=None, credit=2.0):
    r = ScoreRecord(student_id=1, academic_year_id=1, semester=sem, course_code=code,
                    course_name=code, teacher="", credit=credit, score_raw=str(num or ""),
                    score_num=num, gpa=gpa)
    r.id = rid
    return r


def test_collapse_retakes_latest():
    recs = [_rec("C1", "秋季", 55, rid=1), _rec("C1", "春季", 80, rid=2), _rec("C2", "秋季", 90, rid=3)]
    kept = collapse_retakes(recs, "latest")
    assert len(kept) == 2
    c1 = next(r for r in kept if r.course_code == "C1")
    assert c1.score_num == 80  # 取最新学期


def test_collapse_retakes_highest():
    recs = [_rec("C1", "春季", 80, rid=2), _rec("C1", "秋季", 55, rid=1)]
    kept = collapse_retakes(recs, "highest")
    assert kept[0].score_num == 80
    # 无百分制记录参与：只有 NULL 成绩时取最新
    recs2 = [_rec("C9", "秋季", None, rid=1), _rec("C9", "春季", None, rid=2)]
    kept2 = collapse_retakes(recs2, "highest")
    assert len(kept2) == 1 and kept2[0].id == 2


# ---------- 6. 排名（同分同名次 1,1,3 式） ----------
def test_competition_ranking():
    results = []
    for i, (avg, final) in enumerate([(90, 88), (90, 88), (85, 95), (None, 80)]):
        r = StudentResult(student_id=i, student_no=str(i), name=f"s{i}", class_name="c")
        r.weighted_avg, r.final_score = avg, final
        results.append(r)
    compute_rankings(results)
    by_no = {r.student_no: r for r in results}
    assert by_no["0"].academic_rank == 1 and by_no["1"].academic_rank == 1
    assert by_no["2"].academic_rank == 3  # 1,1,3 式
    assert by_no["3"].academic_rank is None  # 无加权平均不参与智育排名
    # 综测排名按综合测评成绩：95 > 88 > 80
    assert by_no["2"].eval_rank == 1
    assert by_no["0"].eval_rank == 2 and by_no["1"].eval_rank == 2
    assert by_no["3"].eval_rank == 4


# ---------- 6b. 专业分组键（v1.3.2 按专业排名） ----------
def test_major_group():
    from app.services.calc import major_group
    assert major_group("建筑类2401") == "建筑类"
    assert major_group("建筑类2402") == "建筑类"       # 同专业不同班 → 同组
    assert major_group("计科2401") == "计科"
    assert major_group("建筑类2402班") == "建筑类"      # 兜底「班」后缀
    assert major_group("城规2101") == "城规"
    assert major_group("试验班") == "试验班"           # 无班号 → 回退全名
    assert major_group("2401") == "2401"
    assert major_group("") == ""
    assert major_group(None) == ""


# ---------- 7. 综测封顶与合成（§5） ----------
def test_eval_cap_and_final():
    # 5 项合计被各自满分封顶：思想品德 26 → 25
    scheme_items = [dict(i) for i in DEFAULT_ITEMS]
    scores = {"思想品德": (26, 25), "社会工作": (21, 20), "科研及科技创新": (10, 10),
              "文体活动": (15, 15), "集体建设": (20, 20)}
    for name, (raw, capped) in scores.items():
        assert min(raw, next(i["max_score"] for i in scheme_items if i["name"] == name)) == capped
    weighted_avg, eval_total = 85.0, 90.0
    final = weighted_avg * 0.8 + eval_total * 0.2
    assert abs(final - 86.0) < 1e-9


# ---------- 8. 2.xlsx 解析（用真实样本） ----------
SAMPLE2 = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                       "..", "..", "2.xlsx"))


@pytest.mark.skipif(not os.path.exists(SAMPLE2), reason="样本 2.xlsx 不存在")
def test_parse_eval_workbook_sample():
    with open(SAMPLE2, "rb") as f:
        parsed = parse_eval_workbook(f.read())
    assert "思想品德" in parsed.item_names
    assert parsed.item_maxes.get("思想品德") == 25
    assert parsed.item_maxes.get("文体活动") == 15
    assert len(parsed.students) == 27  # 样本实际 27 行数据（含 1 名 2102 班学生）
    first = parsed.students[0]
    assert first.student_no.isdigit()
    assert set(first.items.keys()) >= {"思想品德", "社会工作", "科研及科技创新", "文体活动", "集体建设"}
    # 文件自带「总分」应等于五项文件得分之和（首个学生 59.5）
    total = sum(v["score"] for v in first.items.values())
    assert abs(total - 59.5) < 0.01
    # 明细软校验：田鑫媛 文体活动明细求和 = 15（封顶值）
    wd = first.items["文体活动"]
    assert wd["score"] == 15 and wd["soft_sum"] is not None


# ---------- 9. 导出工作簿结构 ----------
def test_export_workbook_structure():
    from openpyxl import Workbook
    from app.services.export import build_grade_sheet
    from app.services.calc import StudentResult
    wb = Workbook()
    ws = wb.active

    class FakeKlass:
        name = "建筑类2401"
    class FakeStudent:
        id = 1
        student_no = "20246631"
        name = "姚思琦"
        klass = FakeKlass()
    res = StudentResult(student_id=1, student_no="20246631", name="姚思琦", class_name="建筑类2401")
    res.weighted_total, res.total_credit, res.weighted_avg = 5795.75, 63.5, 91.27
    res.eval_total, res.final_score = 59.5, 85.0
    res.academic_rank, res.eval_rank = 1, 2
    for it in DEFAULT_ITEMS:
        res.item_scores[it["name"]] = 10.0
        res.items_entered[it["name"]] = True

    class FakeScheme:
        retake_rule = "latest"
        items = DEFAULT_ITEMS
    build_grade_sheet(ws, "24级", [FakeStudent()], "score", FakeScheme(), {1: res}, kept=[])
    assert ws.title == "24级（成绩）"
    assert ws.cell(1, 3).value == "课程代码"
    assert ws.cell(3, 3).value == "学分"
    assert [ws.cell(4, c).value for c in (1, 2, 3)] == ["班级", "学号", "姓名"]
    # 尾部列标签齐全（§13.4），位于标签行（第 4 行）
    row4 = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
    for lab in ["加权总分", "总学分", "加权平均值", "德育", "综合素质测评", "综合测评成绩",
                "智育排名", "综测排名", "确认签字"]:
        assert lab in row4, f"缺少尾部列 {lab}"
    # 加权平均值单元格写入公式
    row5 = [ws.cell(5, c).value for c in range(1, ws.max_column + 1)]
    formulas = [v for v in row5 if isinstance(v, str) and v.startswith("=")]
    assert formulas and formulas[0].startswith("=")

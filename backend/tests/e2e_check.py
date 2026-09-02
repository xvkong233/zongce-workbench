"""端到端验证：登录 → 导入 1.xls → 导入 2.xlsx → 汇总 → 导出工作簿 → 与 output.xlsx 比对。

运行：cd backend && ZONGCE_DATA_DIR=../data-e2e python tests/e2e_check.py
"""
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi.testclient import TestClient  # noqa: E402
from app.main import app  # noqa: E402

client = TestClient(app)


def api(method, path, expect=200, **kw):
    r = client.request(method, f"/api{path}", **kw)
    assert r.status_code == expect, f"{method} {path} → {r.status_code}: {r.text[:400]}"
    return r


# 1. 登录（首次强制改密）
r = api("POST", "/auth/login", json={"username": "admin", "password": "admin123"})
assert r.json()["must_change_password"] is True
tok0 = r.json()["token"]
api("PUT", "/auth/password", json={"old_password": "admin123", "new_password": "zongce-2026"},
    headers={"Authorization": f"Bearer {tok0}"})
r = api("POST", "/auth/login", json={"username": "admin", "password": "zongce-2026"})
token = r.json()["token"]
H = {"Authorization": f"Bearer {token}"}
print("[1] 登录/改密 OK")

# 2. 导入 1.xls：预览
with open(os.path.join(ROOT, "1.xls"), "rb") as f:
    r = api("POST", "/scores/import/preview", headers=H,
            files={"file": ("1.xls", f, "application/vnd.ms-excel")})
pv = r.json()
print(f"[2] 1.xls 预览：学年{pv['years']} 学生{pv['student_count']} 课程{pv['course_count']} "
      f"记录{pv['record_count']} 异常{pv['exception_count']} 冲突{len(pv['conflicts'])}")
assert pv["years"] == ["2024-2025"]
plan = {"create_years": pv["create_years"], "create_grades": pv["create_grades"],
        "create_classes": pv["create_classes"], "conflicts":
        [{"student_no": c["student_no"], "resolve": "system"} for c in pv["conflicts"]]}
with open(os.path.join(ROOT, "1.xls"), "rb") as f:
    r = api("POST", "/scores/import/confirm", headers=H,
            files={"file": ("1.xls", f, "application/vnd.ms-excel")}, data={"plan": __import__("json").dumps(plan)})
print(f"[2] 1.xls 入库：{r.json()['stats']}")

# 3. 综测方案默认存在
scheme = api("GET", "/schemes/default", headers=H).json()
assert len(scheme["items"]) == 5
print("[3] 默认综测方案 OK（五项）")

# 4. 查出学年/年级/班级 id
years = api("GET", "/base/academic-years", headers=H).json()
year_id = next(y["id"] for y in years if y["name"] == "2024-2025")
grades = api("GET", "/base/grades", headers=H).json()
g24 = next(g for g in grades if g["name"] == "24级")
classes = api("GET", "/base/classes", headers=H, params={"grade_id": g24["id"]}).json()
print(f"[4] 基础数据：{len(classes)} 个班级 {[[c['name'] for c in classes]]}")

# 5. 导入 2.xlsx（按年级、学号自动匹配，无需班级映射）
with open(os.path.join(ROOT, "2.xlsx"), "rb") as f:
    r = api("POST", "/evals/import/preview", headers=H,
            params={"academic_year_id": year_id, "grade_id": g24["id"]},
            files={"file": ("2.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
ev = r.json()
print(f"[5] 2.xlsx 预览：项目{ev['item_names']} 学生{ev['student_count']} "
      f"匹配{ev['matched_count']} 未匹配{ev['unmatched_count']} 明细不符{ev['soft_mismatch']} 范围={ev['year']}·{ev['grade']}")
assert set(ev["item_names"]) == {"思想品德", "社会工作", "科研及科技创新", "文体活动", "集体建设"}
assert ev["grade"] == "24级" and ev["matched_count"] == 25, (ev["grade"], ev["matched_count"])
resolve = {u["student_no"]: "file" for u in ev["unmatched"]}
with open(os.path.join(ROOT, "2.xlsx"), "rb") as f:
    r = api("POST", "/evals/import/confirm", headers=H,
            data={"academic_year_id": str(year_id),
                  "grade_id": str(g24["id"]),
                  "resolve": __import__("json").dumps(resolve)},
            files={"file": ("2.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
eval_batch1 = r.json()["batch_id"]
print(f"[5] 2.xlsx 入库：{r.json()['stats']}（批次#{eval_batch1}）")

# 6. 汇总表
r = api("GET", "/summary", headers=H, params={"academic_year_id": year_id, "grade_id": g24["id"]})
sm = r.json()
rows = sm["rows"]
print(f"[6] 汇总：{len(rows)} 人；前 3 名（综测排名）：")
for row in rows[:3]:
    print("   ", row["student_no"], row["name"], "加权平均", row["weighted_avg"],
          "综素", row["eval_total"], "综合测评成绩", row["final_score"], "综测排名", row["eval_rank"])
entered = [r_ for r_ in rows if r_["eval_entered"]]
print(f"    已录入综测 {len(entered)} 人，未录入 {len(rows)-len(entered)} 人（未录入者综素=0）")

# 7. 导出完整工作簿
r = api("POST", "/export/workbook", headers=H,
        json={"academic_year_id": year_id, "grade_ids": [g24["id"]]})
out_path = os.path.join(ROOT, "export_24级.xlsx")
with open(out_path, "wb") as f:
    f.write(r.content)
print(f"[7] 导出工作簿 → {out_path}（{len(r.content)} bytes）")

# 8. 简表导出
r = api("POST", "/export/workbook", headers=H,
        json={"academic_year_id": year_id, "grade_ids": [g24["id"]], "brief": True})
brief_path = os.path.join(ROOT, "export_24级_简表.xlsx")
with open(brief_path, "wb") as f:
    f.write(r.content)
print(f"[8] 简表导出 → {brief_path}")

# ---------- 以下为 v1.3.1 修复项验证 ----------
import io as _io
import json as _json

import openpyxl as _oxl

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# 9. 422 校验错误可读化（§11.10）
r = client.put("/api/auth/password", json={"old_password": "zongce-2026", "new_password": "123"}, headers=H)
assert r.status_code == 422, r.status_code
d = r.json()["detail"]
assert isinstance(d, dict) and "参数校验失败" in d["message"], d
print(f"[9] 422 可读化 OK：{d['message']}")

# 10. 未匹配的 /api 路径返回 404 JSON（而非 SPA 页面）
r = client.get("/api/no/such/api", headers=H)
assert r.status_code == 404 and isinstance(r.json()["detail"], dict), r.text[:200]
print("[10] /api 未匹配路径 404 JSON OK")

# 11. 辅导员成绩导入新建年级自动绑定（§4.1.3）
api("POST", "/users", headers=H,
    json={"username": "c1", "password": "counselor-1", "real_name": "导员甲",
          "enabled": True, "grade_ids": []})
rc = api("POST", "/auth/login", json={"username": "c1", "password": "counselor-1"})
assert rc.json()["must_change_password"] is True
api("PUT", "/auth/password", json={"old_password": "counselor-1", "new_password": "counselor-2"},
    headers={"Authorization": f"Bearer {rc.json()['token']}"})
rc = api("POST", "/auth/login", json={"username": "c1", "password": "counselor-2"})
HC = {"Authorization": f"Bearer {rc.json()['token']}"}

wb = _oxl.Workbook(); ws = wb.active
ws.append(["学年度", "学期", "学号", "姓名", "班级", "课程代码", "课程名称", "授课教师", "学分", "分数", "绩点"])
ws.append(["2026-2027", "秋季", "20276601", "测试生", "建筑类2701", "X001", "测试课程", "师", "2.0", 88, 3.5])
buf = _io.BytesIO(); wb.save(buf)
r = api("POST", "/scores/import/preview", headers=HC, files={"file": ("new.xlsx", buf.getvalue(), XLSX_MIME)})
pv = r.json()
assert pv["create_grades"] and pv["create_grades"][0]["name"] == "27级", pv["create_grades"]
api("POST", "/scores/import/confirm", headers=HC,
    files={"file": ("new.xlsx", buf.getvalue(), XLSX_MIME)},
    data={"plan": _json.dumps({"create_years": pv["create_years"], "create_grades": pv["create_grades"],
                               "create_classes": pv["create_classes"], "conflicts": []})})
my_grades = api("GET", "/base/grades", headers=HC).json()
assert any(g["name"] == "27级" for g in my_grades), my_grades
print("[11] 辅导员导入自动绑定新建年级 OK（绑定 27级）")

# 12. 辅导员向所辖年级之外的已有班级导入名单 → 403
wb = _oxl.Workbook(); ws = wb.active
ws.append(["学号", "姓名", "班级"])
ws.append(["99999001", "越权生", "建筑类2401"])  # 24级不属于 c1
buf = _io.BytesIO(); wb.save(buf)
r = client.post("/api/base/students/import", headers=HC, files={"file": ("x.xlsx", buf.getvalue(), XLSX_MIME)})
assert r.status_code == 403, r.status_code
print("[12] 名单导入越权拦截 OK：", r.json()["detail"]["message"])
# 综测导入按所选年级鉴权：c1 只有 27级，向 24级导入 → 403
with open(os.path.join(ROOT, "2.xlsx"), "rb") as f:
    r = client.post("/api/evals/import/preview", headers=HC,
                    params={"academic_year_id": year_id, "grade_id": g24["id"]},
                    files={"file": ("2.xlsx", f, XLSX_MIME)})
assert r.status_code == 403, r.status_code
print("[12] 综测导入年级鉴权 OK")

# 13. 批次撤销语义：覆盖型批次撤销后恢复旧值与旧批次归属，不再误删记录
with open(os.path.join(ROOT, "2.xlsx"), "rb") as f:
    data2 = f.read()
api("POST", "/evals/import/confirm", headers=H,
    data={"academic_year_id": str(year_id), "grade_id": str(g24["id"]),
          "resolve": _json.dumps(resolve)},
    files={"file": ("2.xlsx", data2, XLSX_MIME)})
from app.database import SessionLocal  # noqa: E402
from app.models import EvalRecord as EvalRecordM  # noqa: E402
r = api("GET", "/batches", headers=H, params={"kind": "eval"}).json()
batch2 = max(b["id"] for b in r["items"] if not b["reverted"])
r = api("POST", f"/batches/{batch2}/revert", headers=H)
print(f"[13] 撤销批次#{batch2}：{r.json()}")
with SessionLocal() as db:
    n_batch2_after = db.query(EvalRecordM).filter(EvalRecordM.batch_id == batch2).count()
    n_batch1_after = db.query(EvalRecordM).filter(EvalRecordM.batch_id == eval_batch1).count()
assert n_batch2_after == 0, f"撤销后仍残留 {n_batch2_after} 条新批次记录"
assert n_batch1_after == 125, f"恢复旧值后 batch1 记录应 125 条，实际 {n_batch1_after}"
print("[13] 批次撤销恢复旧值与旧批次归属 OK")

# 14. 成绩长表样例模板可导入（学生标签列 ↔ 解析器一致）
r = client.get("/api/scores/template", headers=H)
assert r.status_code == 200, r.status_code
tpl = r.content
r = api("POST", "/scores/import/preview", headers=H, files={"file": ("样例.xlsx", tpl, XLSX_MIME)})
pv = r.json()
assert pv["record_count"] == 2, pv["record_count"]
r = api("POST", "/scores/import/confirm", headers=H,
        files={"file": ("样例.xlsx", tpl, XLSX_MIME)},
        data={"plan": _json.dumps({"create_years": pv["create_years"], "create_grades": pv["create_grades"],
                                   "create_classes": pv["create_classes"], "conflicts": []})})
st = r.json()["stats"]
assert st["records_created"] == 2 and st["students_created"] == 1, st
print(f"[14] 成绩样例模板导入 OK：{st}")

# 15. 排名按专业（v1.3.2）：同年级不同专业分别排名，互不影响
SUM_P = {"academic_year_id": year_id, "grade_id": g24["id"]}
before = {r["student_no"]: (r["academic_rank"], r["eval_rank"])
          for r in api("GET", "/summary", headers=H, params=SUM_P).json()["rows"]}
assert len(before) == 51  # 50 名建筑类学生 + §14 模板导入的张三（无 2024-2025 成绩，排名为空）
api("POST", "/base/classes", headers=H, json={"name": "计科2401", "grade_id": g24["id"]})
wb = _oxl.Workbook(); ws = wb.active
ws.append(["学号", "姓名", "班级"])
ws.append(["29999001", "计科霸主", "计科2401"])
buf = _io.BytesIO(); wb.save(buf)
api("POST", "/base/students/import", headers=H, files={"file": ("r.xlsx", buf.getvalue(), XLSX_MIME)})
wb = _oxl.Workbook(); ws = wb.active
ws.append(["学年度", "学期", "学号", "姓名", "班级", "课程代码", "课程名称", "授课教师", "学分", "分数", "绩点"])
ws.append(["2024-2025", "秋季", "29999001", "计科霸主", "计科2401", "TEST999", "测试高分课", "师", "5.0", 100, 5.0])
buf = _io.BytesIO(); wb.save(buf)
r = api("POST", "/scores/import/preview", headers=H, files={"file": ("r.xlsx", buf.getvalue(), XLSX_MIME)})
pv = r.json()
assert not pv["create_years"] and not pv["create_classes"], (pv["create_years"], pv["create_classes"])
api("POST", "/scores/import/confirm", headers=H,
    files={"file": ("r.xlsx", buf.getvalue(), XLSX_MIME)},
    data={"plan": _json.dumps({"create_years": [], "create_grades": [], "create_classes": [], "conflicts": []})})
rows = api("GET", "/summary", headers=H, params=SUM_P).json()["rows"]
after = {r["student_no"]: (r["academic_rank"], r["eval_rank"]) for r in rows}
assert len(after) == 52
for no, ranks in before.items():
    assert after[no] == ranks, f"建筑类学生 {no} 排名被计科学生影响：{ranks} → {after[no]}"
assert after["29999001"] == (1, 1), after["29999001"]  # 计科2401 专业组内双第 1
# 学生报告：排名同样按专业全体计算（而非传入单人所产生的恒 1）
target = next(r for r in rows if r["academic_rank"] == max(
    x["academic_rank"] for x in rows
    if (x["class_name"] == "建筑类2401" or x["class_name"] == "建筑类2402")
    and x["academic_rank"] is not None))
rep = api("GET", "/export/student-report", headers=H,
          params={"student_id": target["student_id"], "academic_year_id": year_id}).json()
assert rep["summary"]["academic_rank"] == target["academic_rank"], (rep["summary"], target)
print(f"[15] 排名按专业 OK：计科2401 双第 1；建筑类 50 人排名不变（末名报告=第 {rep['summary']['academic_rank']}）")

# 15b. 显式专业字段覆盖自动提取：把 计科2401 归入「建筑类」→ 与建筑类一起排名
jk = next(c for c in api("GET", "/base/classes", headers=H,
                         params={"grade_id": g24["id"]}).json() if c["name"] == "计科2401")
assert jk["major"] is None and jk["major_effective"] == "计科", jk
api("PUT", f"/base/classes/{jk['id']}", headers=H,
    json={"name": "计科2401", "grade_id": g24["id"], "major": "建筑类"})
merged = {r["student_no"]: (r["academic_rank"], r["eval_rank"])
          for r in api("GET", "/summary", headers=H, params=SUM_P).json()["rows"]}
prev_top = next(no for no, (ar, _) in after.items() if ar == 1 and no != "29999001")
assert merged["29999001"][0] == 1 and merged["29999001"][1] > 1, merged["29999001"]
# 智育仍第 1；综测（80 分）在合并组内退到多名建筑类学生之后
assert merged[prev_top][0] == 2, (prev_top, merged[prev_top])  # 原建筑类智育第 1 被挤到第 2
api("PUT", f"/base/classes/{jk['id']}", headers=H,
    json={"name": "计科2401", "grade_id": g24["id"], "major": None})
restored = {r["student_no"]: (r["academic_rank"], r["eval_rank"])
            for r in api("GET", "/summary", headers=H, params=SUM_P).json()["rows"]}
assert restored == after, "清除显式专业后应恢复自动分组排名"
print("[15b] 显式专业覆盖/恢复 OK：计科2401 归入建筑类后合并排名，清除后恢复独立分组")

# 15c. 学生报告历年总览（学生管理页「报告」数据源：不带学年参数）
rep = api("GET", "/export/student-report", headers=H,
          params={"student_id": target["student_id"]}).json()
assert "summaries" in rep and len(rep["summaries"]) == 3, rep.get("summaries")
by_year = {x["year"]: x for x in rep["summaries"]}
y24 = by_year["2024-2025"]
assert y24["has_data"] and y24["score_count"] > 0, y24  # 成绩维度（末名学生，2401 班无综测）
assert y24["academic_rank"] == target["academic_rank"] and y24["eval_rank"] == target["eval_rank"]
assert by_year["2026-2027"]["has_data"] is False  # 2026-2027 只有 27级辅导员测试生，target 无数据
# 综测维度：综测已录入学生（20246642，2.xlsx 建筑类2402）应看到 5 条综测记录
top_eval = next(r for r in rows if r["student_no"] == "20246642")
rep2 = api("GET", "/export/student-report", headers=H,
           params={"student_id": top_eval["student_id"]}).json()
y24b = next(x for x in rep2["summaries"] if x["year"] == "2024-2025")
assert y24b["eval_count"] == 5 and y24b["eval_entered"], y24b
print(f"[15c] 学生报告历年总览 OK：3 个学年；成绩维度（末名第 {y24['academic_rank']}）与综测维度（5 条记录）均正确")

# 15d. 学生管理列表：学年维度指标 + 服务端排序
r = api("GET", "/base/students", headers=H, params={
    "academic_year_id": year_id, "page_size": 100,
    "sort": "weighted_avg", "order": "desc"}).json()
vals = [x["weighted_avg"] for x in r["items"] if x["weighted_avg"] is not None]
assert vals == sorted(vals, reverse=True), "weighted_avg 降序排序失败"
assert r["items"][-1]["weighted_avg"] is None, "无成绩学生应排在末尾"
top_eval = next(x for x in r["items"] if x["student_no"] == "20246642")
assert top_eval["eval_entered"] and abs(top_eval["eval_total"] - 77.8) < 0.01, top_eval
r2 = api("GET", "/base/students", headers=H, params={
    "academic_year_id": year_id, "page_size": 100, "sort": "student_no", "order": "asc"}).json()
nos2 = [x["student_no"] for x in r2["items"]]
assert nos2 == sorted(nos2), "学号升序排序失败"
print(f"[15d] 学生列表学年指标与排序 OK：加权平均降序（{len(vals)} 人有值、无成绩者殿后）、综测字段正确")

# 15e. 专业列与专业筛选（学生管理 + 综测汇总）
r = api("GET", "/base/students", headers=H, params={
    "academic_year_id": year_id, "page_size": 100, "major": "建筑类"}).json()
assert r["total"] == 52, r["total"]  # 24级建筑类（51 人）+ 27级建筑类2701 测试生（§11）
assert all(x["major_effective"] == "建筑类" for x in r["items"])
r = api("GET", "/base/students", headers=H, params={
    "academic_year_id": year_id, "page_size": 100, "major": "计科"}).json()
assert r["total"] == 1 and r["items"][0]["student_no"] == "29999001", r["total"]
sm = api("GET", "/summary", headers=H,
         params={"academic_year_id": year_id, "grade_id": g24["id"], "major": "计科"}).json()
assert len(sm["rows"]) == 1 and sm["rows"][0]["student_no"] == "29999001"
sm = api("GET", "/summary", headers=H,
         params={"academic_year_id": year_id, "grade_id": g24["id"], "major": "建筑类"}).json()
assert len(sm["rows"]) == 51, len(sm["rows"])
print("[15e] 专业筛选 OK：学生列表（建筑类 51 / 计科 1）与汇总页（计科 1 / 建筑类 51）")

# 15f. 综测明细样例模板可下载且可被解析器读取
r = client.get("/api/evals/template", headers=H)
assert r.status_code == 200, r.status_code
from app.services.eval_import import parse_eval_workbook  # noqa: E402
tpl_eval = parse_eval_workbook(r.content)
assert tpl_eval.item_names == ["思想品德", "社会工作", "科研及科技创新", "文体活动", "集体建设"], tpl_eval.item_names
assert tpl_eval.item_maxes.get("思想品德") == 25 and len(tpl_eval.students) == 2
first = tpl_eval.students[0]
assert first.items["思想品德"]["score"] == 23 and first.items["思想品德"]["soft_sum"] == 23
print("[15f] 综测样例模板 OK：五项满分识别、2 名样例学生、明细软校验一致")

# 16. 数据清理：按学年+年级清空综测/成绩记录（§10.2）
r = api("POST", "/data/clear", headers=H,
        json={"academic_year_id": year_id, "grade_id": g24["id"], "kind": "eval"})
assert r.json()["deleted"] == 125, r.json()
r = api("POST", "/data/clear", headers=H,
        json={"academic_year_id": year_id, "grade_id": g24["id"], "kind": "score"})
assert r.json()["deleted"] == 1350, r.json()  # 1349 + TEST999
print("[16] 按学年+年级清空数据 OK：综测 125 条、成绩 1350 条")

print("E2E OK")

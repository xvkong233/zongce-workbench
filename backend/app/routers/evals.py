"""综测录入：名册/保存/批量保存、2.xlsx 导入、复制上一学年、批量填充基础分。"""
import json

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session

from ..auth import counselor_grade_ids, get_current_user
from ..database import get_db
from ..models import (AcademicYear, ClassInfo, EvalRecord, Grade, ImportBatch,
                      OperationLog, Student, User)
from ..schemas import CopyPrevIn, EvalBatchIn, EvalSaveIn, FillBaseIn
from ..services.calc import resolve_scheme
from ..services.convert import sum_detail_terms
from ..services.eval_import import confirm_eval_import, parse_eval_workbook

router = APIRouter(prefix="/evals", tags=["evals"])


@router.get("/template")
def eval_template(_: User = Depends(get_current_user)):
    """下载综测明细样例表格（双层表头：项目大类合并单元格跨「加减分项/得分」两列）。"""
    import io
    import openpyxl
    from fastapi import Response
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote

    from ..services.calc import DEFAULT_ITEMS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "综测明细"
    ws.cell(1, 1, "班级")
    ws.cell(1, 2, "学号")
    ws.cell(1, 3, "姓名")
    col = 4
    for item in DEFAULT_ITEMS:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(1, col, f"{item['name']}{item['max_score']:.0f}")
        col += 2
    ws.cell(1, col, "总分")

    samples = [
        ("建筑类2401", "20246631", "张三",
         [("基础分+23", 23), ("部长+2", 2), ("专利+5", 5), ("院运会+3", 3),
          ("班级基础分+7\n寝室基础分+8", 15)], 68),
        ("建筑类2401", "20246632", "李四",
         [("基础分+20", 20), ("", 0), ("", 0), ("班级篮球赛+1", 1), ("", 0)], 21),
    ]
    for r, (klass, no, name, items, total) in enumerate(samples, start=2):
        ws.cell(r, 1, klass)
        ws.cell(r, 2, no)
        ws.cell(r, 3, name)
        c = 4
        for detail, score in items:
            ws.cell(r, c, detail)
            ws.cell(r, c + 1, score)
            c += 2
        ws.cell(r, c, total)

    for c in range(1, col + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('综测明细样例.xlsx')}"},
    )


def _class_and_access(db: Session, user: User, class_id: int) -> ClassInfo:
    klass = db.get(ClassInfo, class_id)
    if not klass:
        raise HTTPException(404, {"message": "班级不存在"})
    allowed = counselor_grade_ids(user)
    if allowed is not None and klass.grade_id not in allowed:
        raise HTTPException(403, {"message": "无权访问该班级"})
    return klass


@router.get("/roster")
def roster(academic_year_id: int, class_id: int, db: Session = Depends(get_db),
           user: User = Depends(get_current_user)):
    klass = _class_and_access(db, user, class_id)
    scheme = resolve_scheme(db, academic_year_id, klass.grade_id)
    students = db.query(Student).filter_by(class_id=class_id).order_by(Student.student_no).all()
    records = db.query(EvalRecord).filter_by(academic_year_id=academic_year_id).filter(
        EvalRecord.student_id.in_([s.id for s in students] or [0])).all()
    by_student: dict[int, dict] = {}
    for r in records:
        soft = sum_detail_terms(r.detail_text) if r.detail_text else None
        by_student.setdefault(r.student_id, {})[r.item_name] = {
            "detail_text": r.detail_text, "score": r.score, "soft_sum": soft,
            "mismatch": bool(soft is not None and abs(soft - r.score) > 0.05)}
    items = [{"name": i["name"], "max_score": i["max_score"],
              "base_template": i.get("base_template", "")} for i in scheme.items]
    rows = []
    for s in students:
        item_map = by_student.get(s.id, {})
        entered = any(it.get("detail_text") or it.get("score") for it in item_map.values())
        subtotal = 0.0
        item_cells = []
        for it in items:
            data = item_map.get(it["name"])
            score = min(data["score"], it["max_score"]) if data else None
            if score is not None:
                subtotal += score
            item_cells.append({**(data or {}), "capped": score})
        rows.append({"student_id": s.id, "student_no": s.student_no, "name": s.name,
                     "entered": entered, "items": item_cells, "subtotal": round(subtotal, 2)})
    return {"items": items, "rows": rows, "weight_academic": scheme.weight_academic,
            "weight_eval": scheme.weight_eval}


def _save_eval_records(db: Session, user: User, body: EvalSaveIn) -> list[str]:
    """写入单个学生的综测记录（不提交事务），返回明细不符的项目名。"""
    s = db.get(Student, body.student_id)
    if not s:
        raise HTTPException(404, {"message": "学生不存在"})
    _class_and_access(db, user, s.class_id)
    mismatches = []
    for it in body.items:
        soft = sum_detail_terms(it.detail_text) if it.detail_text else None
        if soft is not None and abs(soft - it.score) > 0.05:
            mismatches.append(it.item_name)
        rec = db.query(EvalRecord).filter_by(student_id=body.student_id,
                                             academic_year_id=body.academic_year_id,
                                             item_name=it.item_name).first()
        if rec:
            rec.detail_text, rec.score = it.detail_text, it.score
        else:
            db.add(EvalRecord(student_id=body.student_id, academic_year_id=body.academic_year_id,
                              item_name=it.item_name, detail_text=it.detail_text, score=it.score))
    return mismatches


@router.put("/save")
def save_eval(body: EvalSaveIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    mismatches = _save_eval_records(db, user, body)
    db.commit()
    return {"ok": True, "mismatches": mismatches}


@router.post("/batch")
def batch_save(body: EvalBatchIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """全班保存：单事务提交，任一学生失败则整体回滚。"""
    for row in body.rows:
        _save_eval_records(db, user, EvalSaveIn(student_id=row.student_id,
                                                academic_year_id=body.academic_year_id,
                                                items=row.items))
    db.commit()
    return {"ok": True, "count": len(body.rows)}


@router.post("/import/preview")
def eval_import_preview(file: UploadFile = File(...), academic_year_id: int = 0,
                        grade_id: int = 0,
                        db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """按年级导入：仅以学号匹配所选年级内的学生，文件班级列仅作参考、无需映射。"""
    try:
        parsed = parse_eval_workbook(file.file.read())
    except ValueError as e:
        raise HTTPException(400, {"message": str(e)})
    allowed = counselor_grade_ids(user)
    if grade_id <= 0:
        raise HTTPException(400, {"message": "请先选择目标年级"})
    if allowed is not None and grade_id not in allowed:
        raise HTTPException(403, {"message": "无权向该年级导入"})
    grade = db.get(Grade, grade_id)
    if not grade:
        raise HTTPException(400, {"message": "目标年级不存在"})
    year = db.get(AcademicYear, academic_year_id)
    if not year:
        raise HTTPException(400, {"message": "请先选择目标学年"})
    nos = [s.student_no for s in parsed.students]
    found = set()
    if nos:
        found = {s.student_no for s in db.query(Student).join(ClassInfo).filter(
            ClassInfo.grade_id == grade_id,
            Student.student_no.in_(nos)).all()}
    unmatched = [{"student_no": s.student_no, "name": s.name, "class": s.class_name_raw}
                 for s in parsed.students if s.student_no not in found]
    soft_mismatch = 0
    for s in parsed.students:
        for item in s.items.values():
            if item.get("soft_sum") is not None and item.get("score") is not None \
               and abs(item["soft_sum"] - item["score"]) > 0.05:
                soft_mismatch += 1
    return {"filename": file.filename, "title": parsed.title, "year": year.name,
            "grade": grade.name,
            "item_names": parsed.item_names, "item_maxes": parsed.item_maxes,
            "class_keys": parsed.class_keys,
            "student_count": len(parsed.students),
            "matched_count": len(found),
            "unmatched": unmatched[:200], "unmatched_count": len(unmatched),
            "soft_mismatch": soft_mismatch}


@router.post("/import/confirm")
def eval_import_confirm(file: UploadFile = File(...), academic_year_id: int = Form(default=0),
                        grade_id: int = Form(default=0),
                        resolve: str = Form(default="{}"),
                        db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    try:
        resolve_obj = json.loads(resolve)
    except json.JSONDecodeError:
        raise HTTPException(400, {"message": "resolve 参数不合法"})
    if grade_id <= 0:
        raise HTTPException(400, {"message": "请先选择目标年级"})
    allowed = counselor_grade_ids(user)
    if allowed is not None and grade_id not in allowed:
        raise HTTPException(403, {"message": "无权向该年级导入"})
    grade = db.get(Grade, grade_id)
    if not grade:
        raise HTTPException(400, {"message": "目标年级不存在"})
    year = db.get(AcademicYear, academic_year_id)
    if not year:
        raise HTTPException(400, {"message": "目标学年不存在，请先选择学年"})
    try:
        parsed = parse_eval_workbook(file.file.read())
    except ValueError as e:
        raise HTTPException(400, {"message": str(e)})
    batch = confirm_eval_import(db, parsed, academic_year_id, grade_id, resolve_obj,
                                user.id, file.filename)
    return {"batch_id": batch.id, "stats": {k: v for k, v in batch.stats.items() if k != "unmatched_list"}}


@router.post("/copy-prev")
def copy_prev(body: CopyPrevIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """复制上一学年：仅填充目标学年未录入的学生，按目标方案映射项目名。"""
    if not body.class_id:
        raise HTTPException(400, {"message": "请选择班级"})
    klass = _class_and_access(db, user, body.class_id)
    students = db.query(Student).filter_by(class_id=body.class_id).order_by(Student.student_no).all()
    target_scheme = resolve_scheme(db, body.to_year_id, klass.grade_id if klass else None)
    target_names = [i["name"] for i in target_scheme.items]
    src = db.query(EvalRecord).filter_by(academic_year_id=body.from_year_id).filter(
        EvalRecord.student_id.in_([s.id for s in students] or [0])).all()
    existing = {(e.student_id, e.item_name) for e in db.query(EvalRecord).filter_by(
        academic_year_id=body.to_year_id).filter(
        EvalRecord.student_id.in_([s.id for s in students] or [0])).all()}
    entered_targets = {sid for sid, _ in existing}
    copied, skipped_students, skipped_items = 0, 0, 0
    by_student_src: dict[int, list] = {}
    for e in src:
        by_student_src.setdefault(e.student_id, []).append(e)
    for s in students:
        if s.id in entered_targets:
            skipped_students += 1
            continue
        for e in by_student_src.get(s.id, []):
            if e.item_name in target_names:
                name = e.item_name
            else:
                match = next((t for t in target_names if t and (t in e.item_name or e.item_name in t)), None)
                if match is None:
                    skipped_items += 1
                    continue
                name = match
            if (s.id, name) in existing:
                continue
            db.add(EvalRecord(student_id=s.id, academic_year_id=body.to_year_id,
                              item_name=name, detail_text=e.detail_text, score=e.score))
            copied += 1
    db.commit()
    return {"copied_records": copied, "skipped_students": skipped_students,
            "skipped_items": skipped_items}


@router.post("/fill-base")
def fill_base(body: FillBaseIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """批量填充基础分模板：按方案模板自动换算得分。"""
    klass = _class_and_access(db, user, body.class_id)
    scheme = resolve_scheme(db, body.academic_year_id, klass.grade_id)
    students = db.query(Student).filter_by(class_id=body.class_id).order_by(Student.student_no).all()
    existing = set()
    if body.only_missing:
        have = db.query(EvalRecord).filter_by(academic_year_id=body.academic_year_id).filter(
            EvalRecord.student_id.in_([s.id for s in students] or [0])).all()
        existing = {e.student_id for e in have}
    filled = 0
    for s in students:
        if s.id in existing:
            continue
        for item in scheme.items:
            template = item.get("base_template") or ""
            if not template:
                continue
            score = sum_detail_terms(template)
            db.add(EvalRecord(student_id=s.id, academic_year_id=body.academic_year_id,
                              item_name=item["name"], detail_text=template, score=score))
            filled += 1
    db.commit()
    return {"filled_records": filled, "students_affected": len(students) - len(existing)}

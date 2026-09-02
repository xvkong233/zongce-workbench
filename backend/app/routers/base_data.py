"""基础数据：学院 / 学年 / 年级 / 班级 / 学生。"""
import io
import re

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from ..auth import counselor_grade_ids, get_current_user, require_admin
from ..database import get_db
from ..models import (AcademicYear, ClassInfo, College, Grade,
                      OperationLog, Student, User)
from ..schemas import ClassIn, StudentIn
from ..services.score_import import _as_text

router = APIRouter(prefix="/base", tags=["base"])


def _log(db, user, action, detail):
    db.add(OperationLog(operator_id=user.id, operator_name=user.username,
                        action=action, detail=detail))


# ---------- 学院 ----------
@router.get("/colleges")
def list_colleges(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return [{"id": c.id, "name": c.name} for c in db.query(College).order_by(College.name).all()]


@router.post("/colleges")
def create_college(body: dict, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, {"message": "学院名称不能为空"})
    if db.query(College).filter_by(name=name).first():
        raise HTTPException(400, {"message": "学院已存在"})
    c = College(name=name)
    db.add(c)
    _log(db, user, "新建学院", name)
    db.commit()
    db.refresh(c)
    return {"id": c.id, "name": c.name}


@router.delete("/colleges/{college_id}")
def delete_college(college_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    c = db.get(College, college_id)
    if not c:
        raise HTTPException(404, {"message": "学院不存在"})
    if db.query(ClassInfo).filter_by(college_id=college_id).count():
        raise HTTPException(400, {"message": "仍有班级隶属该学院，无法删除"})
    _log(db, user, "删除学院", c.name)
    db.delete(c)
    db.commit()
    return {"ok": True}


# ---------- 学年 ----------
@router.get("/academic-years")
def list_years(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return [{"id": y.id, "name": y.name} for y in
            db.query(AcademicYear).order_by(AcademicYear.name.desc()).all()]


def _validate_year_name(name: str):
    m = re.fullmatch(r"(\d{4})-(\d{4})", name.strip())
    if not m or int(m.group(2)) != int(m.group(1)) + 1:
        raise HTTPException(400, {"message": "学年格式应为 YYYY-YYYY 且结束年=开始年+1，如 2024-2025"})


@router.post("/academic-years")
def create_year(body: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    name = (body.get("name") or "").strip()
    _validate_year_name(name)
    if db.query(AcademicYear).filter_by(name=name).first():
        raise HTTPException(400, {"message": "学年已存在"})
    y = AcademicYear(name=name)
    db.add(y)
    _log(db, user, "新建学年", name)
    db.commit()
    db.refresh(y)
    return {"id": y.id, "name": y.name}


@router.put("/academic-years/{year_id}")
def rename_year(year_id: int, body: dict, db: Session = Depends(get_db),
                user: User = Depends(require_admin)):
    y = db.get(AcademicYear, year_id)
    if not y:
        raise HTTPException(404, {"message": "学年不存在"})
    name = (body.get("name") or "").strip()
    _validate_year_name(name)
    old = y.name
    y.name = name
    _log(db, user, "学年改名", f"{old} → {name}")
    db.commit()
    return {"id": y.id, "name": y.name}


@router.delete("/academic-years/{year_id}")
def delete_year(year_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    from ..models import EvalRecord, ScoreRecord
    y = db.get(AcademicYear, year_id)
    if not y:
        raise HTTPException(404, {"message": "学年不存在"})
    if db.query(ScoreRecord).filter_by(academic_year_id=year_id).count() or \
       db.query(EvalRecord).filter_by(academic_year_id=year_id).count():
        raise HTTPException(400, {"message": "该学年已有成绩或综测数据，请先清空相关数据"})
    _log(db, user, "删除学年", y.name)
    db.delete(y)
    db.commit()
    return {"ok": True}


# ---------- 年级 ----------
@router.get("/grades")
def list_grades(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    allowed = counselor_grade_ids(user)
    q = db.query(Grade).order_by(Grade.enrollment_year.desc())
    items = [{"id": g.id, "name": g.name, "enrollment_year": g.enrollment_year} for g in q]
    if allowed is not None:
        items = [i for i in items if i["id"] in allowed]
    return items


@router.post("/grades")
def create_grade(body: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    name = (body.get("name") or "").strip()
    if not re.fullmatch(r"\d{2}级", name):
        raise HTTPException(400, {"message": "年级名称格式应为「XX级」，如 24级"})
    if db.query(Grade).filter_by(name=name).first():
        raise HTTPException(400, {"message": "年级已存在"})
    enrollment = body.get("enrollment_year") or (2000 + int(name[:2]))
    g = Grade(name=name, enrollment_year=enrollment)
    db.add(g)
    if user.role == "counselor":  # 辅导员新建年级自动绑定自己
        user.grades.append(g)
    _log(db, user, "新建年级", name)
    db.commit()
    db.refresh(g)
    return {"id": g.id, "name": g.name, "enrollment_year": g.enrollment_year}


@router.delete("/grades/{grade_id}")
def delete_grade(grade_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    from ..models import ClassInfo
    g = db.get(Grade, grade_id)
    if not g:
        raise HTTPException(404, {"message": "年级不存在"})
    if db.query(ClassInfo).filter_by(grade_id=grade_id).count():
        raise HTTPException(400, {"message": "该年级下仍有班级，无法删除"})
    _log(db, user, "删除年级", g.name)
    db.delete(g)
    db.commit()
    return {"ok": True}


# ---------- 班级 ----------
def _class_item(c: ClassInfo):
    from ..services.calc import major_group
    effective = (c.major or "").strip() or major_group(c.name)
    return {"id": c.id, "name": c.name, "grade_id": c.grade_id, "grade_name": c.grade.name,
            "college_id": c.college_id, "college_name": c.college.name if c.college else None,
            "major": c.major, "major_effective": effective}


@router.get("/classes")
def list_classes(grade_id: int | None = None, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    allowed = counselor_grade_ids(user)
    q = db.query(ClassInfo)
    if grade_id:
        if allowed is not None and grade_id not in allowed:
            raise HTTPException(403, {"message": "无权访问该年级"})
        q = q.filter_by(grade_id=grade_id)
    elif allowed is not None:
        q = q.filter(ClassInfo.grade_id.in_(allowed))
    return [_class_item(c) for c in q.order_by(ClassInfo.name).all()]


@router.post("/classes")
def create_class(body: ClassIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, {"message": "班级名称不能为空"})
    if db.query(ClassInfo).filter_by(name=name).first():
        raise HTTPException(400, {"message": "班级已存在"})
    grade = db.get(Grade, body.grade_id)
    if not grade:
        raise HTTPException(400, {"message": "年级不存在"})
    allowed = counselor_grade_ids(user)
    if allowed is not None and grade.id not in allowed:
        raise HTTPException(403, {"message": "无权在该年级下新建班级"})
    c = ClassInfo(name=name, grade_id=grade.id, college_id=body.college_id,
                  major=(body.major or "").strip() or None)
    db.add(c)
    _log(db, user, "新建班级", name)
    db.commit()
    db.refresh(c)
    return _class_item(c)


@router.put("/classes/{class_id}")
def update_class(class_id: int, body: ClassIn, db: Session = Depends(get_db),
                 user: User = Depends(require_admin)):
    c = db.get(ClassInfo, class_id)
    if not c:
        raise HTTPException(404, {"message": "班级不存在"})
    old = c.name
    c.name = body.name.strip()
    c.grade_id = body.grade_id
    c.college_id = body.college_id
    c.major = (body.major or "").strip() or None
    _log(db, user, "修改班级", f"{old} → {c.name}")
    db.commit()
    return _class_item(c)


@router.delete("/classes/{class_id}")
def delete_class(class_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    from ..models import Student
    c = db.get(ClassInfo, class_id)
    if not c:
        raise HTTPException(404, {"message": "班级不存在"})
    if db.query(Student).filter_by(class_id=class_id).count():
        raise HTTPException(400, {"message": "该班级下仍有学生，无法删除"})
    _log(db, user, "删除班级", c.name)
    db.delete(c)
    db.commit()
    return {"ok": True}


# ---------- 学生 ----------
STUDENT_SORTS = ("student_no", "name", "class_name", "grade_name", "major",
                 "weighted_avg", "avg_gpa", "eval_total", "final_score")


@router.get("/students")
def list_students(class_id: int | None = None, keyword: str = "", page: int = 1,
                  page_size: int = 20, academic_year_id: int | None = None,
                  major: str = "", sort: str = "student_no", order: str = "asc",
                  db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """学生分页查询。带 academic_year_id 时附带该学年维度指标
    （学业加权平均 / GPA / 综测 / 综合测评成绩）；排序在服务端完成（指标列为计算值）。
    major 按班级「有效专业」（显式专业优先，否则按班名自动提取）筛选。"""
    from ..services.calc import compute_metrics_bulk, major_group
    allowed = counselor_grade_ids(user)
    q = db.query(Student)
    if class_id:
        q = q.filter_by(class_id=class_id)
    elif allowed is not None:
        q = q.join(ClassInfo).filter(ClassInfo.grade_id.in_(allowed))
    if keyword:
        q = q.filter((Student.student_no.like(f"%{keyword}%")) | (Student.name.like(f"%{keyword}%")))

    # 班级有效专业映射（班级表很小，全量加载；筛选 / 字段 / 排序共用）
    eff_major = {c.id: ((c.major or "").strip() or major_group(c.name))
                 for c in db.query(ClassInfo).all()}
    if major:
        ids = [cid for cid, m in eff_major.items() if m == major]
        q = q.filter(Student.class_id.in_(ids or [0]))

    total = q.count()
    students = q.all()

    metrics = compute_metrics_bulk(db, students, academic_year_id) \
        if academic_year_id and students else {}
    if sort not in STUDENT_SORTS:
        sort = "student_no"
    desc = order == "desc"

    def key_of(s: Student):
        if sort == "student_no":
            return s.student_no
        if sort == "name":
            return s.name
        if sort == "class_name":
            return s.klass.name if s.klass else ""
        if sort == "grade_name":
            return s.klass.grade.name if s.klass and s.klass.grade else ""
        if sort == "major":
            return eff_major.get(s.class_id, "")
        return metrics.get(s.id, {}).get(sort)

    if sort in ("student_no", "name", "class_name", "grade_name", "major"):
        students.sort(key=key_of, reverse=desc)
    else:  # 指标列：无值者恒排末尾
        with_v = [s for s in students if key_of(s) is not None]
        without_v = [s for s in students if key_of(s) is None]
        with_v.sort(key=key_of, reverse=desc)
        students = with_v + without_v

    start = (page - 1) * page_size
    rows = students[start:start + page_size]
    return {"total": total, "items": [
        {"id": s.id, "student_no": s.student_no, "name": s.name,
         "class_id": s.class_id, "class_name": s.klass.name if s.klass else "",
         "grade_name": s.klass.grade.name if s.klass and s.klass.grade else "",
         "major": s.klass.major if s.klass else None,
         "major_effective": eff_major.get(s.class_id, ""),
         "weighted_avg": metrics.get(s.id, {}).get("weighted_avg"),
         "avg_gpa": metrics.get(s.id, {}).get("avg_gpa"),
         "eval_total": metrics.get(s.id, {}).get("eval_total"),
         "eval_entered": metrics.get(s.id, {}).get("eval_entered", False),
         "final_score": metrics.get(s.id, {}).get("final_score")}
        for s in rows]}


@router.put("/students/{student_id}")
def update_student(student_id: int, body: StudentIn, db: Session = Depends(get_db),
                   user: User = Depends(get_current_user)):
    s = db.get(Student, student_id)
    if not s:
        raise HTTPException(404, {"message": "学生不存在"})
    allowed = counselor_grade_ids(user)
    if allowed is not None and (not s.klass or s.klass.grade_id not in allowed):
        raise HTTPException(403, {"message": "无权修改该学生"})
    dup = db.query(Student).filter(Student.student_no == body.student_no, Student.id != student_id).first()
    if dup:
        raise HTTPException(400, {"message": f"学号 {body.student_no} 已存在（{dup.name}）"})
    old = f"{s.student_no} {s.name} {s.klass.name if s.klass else ''}"
    s.student_no, s.name, s.class_id = body.student_no.strip(), body.name.strip(), body.class_id
    _log(db, user, "修正学籍", f"{old} → {s.student_no} {s.name}")
    db.commit()
    return {"ok": True}


@router.delete("/students/{student_id}")
def delete_student(student_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    from ..models import EvalRecord, ScoreRecord
    s = db.get(Student, student_id)
    if not s:
        raise HTTPException(404, {"message": "学生不存在"})
    detail = f"{s.student_no} {s.name}（级联删除其成绩与综测记录）"
    db.query(ScoreRecord).filter_by(student_id=s.id).delete()
    db.query(EvalRecord).filter_by(student_id=s.id).delete()
    db.delete(s)
    _log(db, user, "删除学生", detail)
    db.commit()
    return {"ok": True}


@router.get("/students/template")
def student_template(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生名单"
    ws.append(["学号", "姓名", "班级"])
    ws.append(["20246631", "张三", "建筑类2401"])
    buf = io.BytesIO()
    wb.save(buf)
    from fastapi import Response
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=student_template.xlsx"})


@router.post("/students/import")
def import_students(file: UploadFile = File(...), db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    """学生名单独立导入（学号/姓名/班级）。班级不存在时自动按班级名建年级+班级；
    班级已存在时校验辅导员年级权限；姓名/班级与系统不一致计入 conflicts 返回。"""
    from ..services.score_import import (infer_enrollment_year, infer_grade_name,
                                         normalize_class_name)
    allowed = counselor_grade_ids(user)
    data = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    created_students, updated_students = 0, 0
    created_classes, errors, conflicts = [], [], []
    for ws in wb.worksheets:
        rows = [[_norm(c.value) for c in r] for r in ws.iter_rows()]
        header_idx = None
        col = {}
        for i, row in enumerate(rows):
            texts = {_as_text(v): j for j, v in enumerate(row) if v is not None}
            if "学号" in texts and "姓名" in texts:
                header_idx = i
                col = {"student_no": texts["学号"],
                       "name": texts["姓名"],
                       "class_name": texts.get("班级")}
                break
        if header_idx is None:
            continue
        for ri in range(header_idx + 1, len(rows)):
            row = rows[ri]
            no = _as_text(row[col["student_no"]]) if col["student_no"] < len(row) else ""
            name = _as_text(row[col["name"]]) if col["name"] < len(row) else ""
            raw_class = _as_text(row[col["class_name"]]) if col.get("class_name") is not None and col["class_name"] < len(row) else ""
            if not no:
                continue
            class_name = normalize_class_name(raw_class)
            if not class_name:
                errors.append(f"第{ri+1}行 学号{no} 缺班级")
                continue
            klass = db.query(ClassInfo).filter_by(name=class_name).first()
            if klass is None:
                grade_name = infer_grade_name(class_name)
                if not grade_name:
                    errors.append(f"第{ri+1}行 班级「{raw_class}」无法推断年级")
                    continue
                grade = db.query(Grade).filter_by(name=grade_name).first()
                if grade is None:
                    grade = Grade(name=grade_name, enrollment_year=infer_enrollment_year(grade_name))
                    if user.role == "counselor":
                        user.grades.append(grade)
                    db.add(grade)
                    db.flush()
                klass = ClassInfo(name=class_name, grade_id=grade.id)
                db.add(klass)
                db.flush()
                created_classes.append(class_name)
            elif allowed is not None and klass.grade_id not in allowed:
                raise HTTPException(403, {"message":
                    f"第{ri+1}行 班级「{klass.name}」属于所辖年级之外的年级，无权导入该学生（学号 {no}）"})
            s = db.query(Student).filter_by(student_no=no).first()
            if s is None:
                db.add(Student(student_no=no, name=name or no, class_id=klass.id))
                created_students += 1
            else:
                diffs = []
                if name and s.name != name:
                    diffs.append(f"姓名 系统「{s.name}」→ 文件「{name}」")
                if s.class_id != klass.id:
                    diffs.append(f"班级 系统「{s.klass.name if s.klass else '—'}」→ 文件「{class_name}」")
                if diffs:
                    conflicts.append({"student_no": no, "change": "；".join(diffs)})
                s.name = name or s.name
                s.class_id = klass.id
                updated_students += 1
    _log(db, user, "学生名单导入",
         f"{file.filename}：新建{created_students} 更新{updated_students} 新建班级{len(created_classes)}"
         + (f" 学籍冲突{len(conflicts)}" if conflicts else ""))
    db.commit()
    return {"created_students": created_students, "updated_students": updated_students,
            "created_classes": created_classes, "errors": errors, "conflicts": conflicts}


def _norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() if isinstance(v, str) else v

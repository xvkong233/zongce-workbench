"""成绩长表导入：预览（解析+异常清单）/ 确认入库 / 学生成绩明细查询 / 异常导出 CSV / 样例表格下载。支持教务新旧两种导出格式（自动识别）。"""
import csv
import io

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..auth import counselor_grade_ids, get_current_user
from ..database import get_db
from ..models import (AcademicYear, ClassInfo, Grade, GradeConversion, Student,
                      User)
from ..services.score_import import (confirm_score_import, find_conflicts_with_db,
                                     infer_enrollment_year, infer_grade_name,
                                     parse_score_workbook)
from .base_data import _class_item

router = APIRouter(prefix="/scores", tags=["scores"])


def _conversion_map(db: Session) -> dict[str, float]:
    return {c.level_text: c.score for c in db.query(GradeConversion).all()}


def _check_grade_access(user: User, class_names: list[str], db: Session):
    allowed = counselor_grade_ids(user)
    if allowed is None:
        return
    for name in class_names:
        klass = db.query(ClassInfo).filter_by(name=name).first()
        if klass and klass.grade_id not in allowed:
            raise HTTPException(403, {"message": f"文件包含所辖年级之外的班级「{name}」，无权导入"})


@router.post("/import/preview")
def import_preview(file: UploadFile = File(...), db: Session = Depends(get_db),
                   user: User = Depends(get_current_user)):
    try:
        parsed = parse_score_workbook(file.filename, file.file.read(), _conversion_map(db))
    except ValueError as e:
        raise HTTPException(400, {"message": str(e)})

    _check_grade_access(user, parsed.class_names, db)

    existing_years = {y.name for y in db.query(AcademicYear).all()}
    existing_grades = {g.name for g in db.query(Grade).all()}
    existing_classes = {c.name for c in db.query(ClassInfo).all()}

    create_years = [{"name": y} for y in parsed.year_names if y not in existing_years]
    class_grade: dict[str, str] = {}
    create_grades, create_classes = [], []
    for cn in parsed.class_names:
        gname = infer_grade_name(cn)
        class_grade[cn] = gname or ""
        if gname and gname not in existing_grades and gname not in {g["name"] for g in create_grades}:
            create_grades.append({"name": gname, "enrollment_year": infer_enrollment_year(gname)})
        if cn not in existing_classes:
            create_classes.append({"name": cn, "grade_name": gname or "", "college_name": None})

    conflicts = find_conflicts_with_db(db, parsed)
    return {
        "filename": file.filename,
        "years": parsed.year_names,
        "class_grade": class_grade,
        "student_count": parsed.student_count,
        "course_count": parsed.course_count,
        "record_count": len(parsed.rows),
        "create_years": create_years,
        "create_grades": create_grades,
        "create_classes": create_classes,
        "exceptions": parsed.exceptions[:500],
        "exception_count": len(parsed.exceptions),
        "conflicts": conflicts,
    }


@router.post("/import/confirm")
def import_confirm(file: UploadFile = File(...), plan: str = Form(default="{}"),
                   db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    import json
    try:
        plan_obj = json.loads(plan)
    except json.JSONDecodeError:
        raise HTTPException(400, {"message": "plan 参数不是合法 JSON"})
    try:
        parsed = parse_score_workbook(file.filename, file.file.read(), _conversion_map(db))
    except ValueError as e:
        raise HTTPException(400, {"message": str(e)})
    _check_grade_access(user, parsed.class_names, db)
    batch = confirm_score_import(db, parsed, plan_obj, user, file.filename)
    return {"batch_id": batch.id, "stats": batch.stats}


@router.get("/records")
def student_records(student_id: int, academic_year_id: int | None = None,
                    db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from ..models import ScoreRecord
    s = db.get(Student, student_id)
    if not s:
        raise HTTPException(404, {"message": "学生不存在"})
    allowed = counselor_grade_ids(user)
    if allowed is not None and (not s.klass or s.klass.grade_id not in allowed):
        raise HTTPException(403, {"message": "无权查看该学生"})
    q = db.query(ScoreRecord).filter_by(student_id=student_id)
    if academic_year_id:
        q = q.filter_by(academic_year_id=academic_year_id)
    years = {y.id: y.name for y in db.query(AcademicYear).all()}
    return [{"id": r.id, "year": years.get(r.academic_year_id, ""), "semester": r.semester,
             "course_code": r.course_code, "course_name": r.course_name, "teacher": r.teacher,
             "credit": r.credit, "score_raw": r.score_raw, "score_num": r.score_num,
             "gpa": r.gpa} for r in q.order_by(ScoreRecord.course_code).all()]


@router.post("/exceptions/export")
async def exceptions_export(exceptions: list[dict],
                            db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["sheet", "行号", "异常类型", "详情"])
    for e in exceptions:
        writer.writerow([e.get("sheet", ""), e.get("row", ""), e.get("type", ""), e.get("detail", "")])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue().encode("utf-8-sig")]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=exceptions.csv"})


@router.get("/template")
def score_template(_: User = Depends(get_current_user)):
    """下载成绩长表样例（与教务最新导出格式一致）。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生成绩"
    headers = ["学号", "姓名", "学生标签", "课程名", "课程号", "学分", "总成绩",
               "平时成绩", "期中成绩", "期末成绩", "其他成绩1", "加分", "绩点",
               "等级成绩", "成绩获得学年学期", "是否参与绩点计算", "显示总成绩", "重修重考"]
    ws.append(headers)
    ws.append(["20246601", "张三", "建筑类2401", "高等数学①㈠", "A1501000015", "5.0",
               89, "88.0", "", "90", "", "", 3.9, "", "2025-2026学年秋", "是", 89, "初修"])
    ws.append(["20246601", "张三", "建筑类2401", "在线公共选修课示例", "A3201001010", "1.0",
               "合格", "", "", "", "", "", 4.0, "合格", "2025-2026学年秋", "否", "合格", "初修"])
    from openpyxl.styles import Alignment, Font, PatternFill
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    from urllib.parse import quote
    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('成绩长表样例.xlsx')}"},
    )

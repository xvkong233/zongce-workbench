"""成绩长表解析与入库：兼容教务旧格式（学年度/学期/课程代码）与新格式（课程号/总成绩/成绩获得学年学期），自动识别。"""
from __future__ import annotations

import io
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field

import openpyxl
import xlrd
from sqlalchemy.orm import Session

from ..models import (AcademicYear, ClassInfo, Grade, ImportBatch,
                      OperationLog, ScoreRecord, Student)
from .convert import convert_level, parse_number

SEMESTER_ALIASES = {"秋": "秋季", "秋季": "秋季", "1": "秋季", "一": "秋季", "上": "秋季",
                    "春": "春季", "春季": "春季", "2": "春季", "二": "春季", "下": "春季"}

# 中文表头 → 内部字段（学生标签=新教务导出的班级列，与「班级」任选其一）
HEADER_MAP = {"学号": "student_no", "课程代码": "course_code", "课程号": "course_code",
              "学年度": "year", "学期": "semester", "成绩获得学年学期": "year_term",
              "姓名": "name", "班级": "class_name", "学生标签": "class_name",
              "课程序号": "course_seq",
              "课程名称": "course_name", "课程名": "course_name", "授课教师": "teacher",
              "课程学分": "credit", "学分": "credit", "总成绩": "score",
              "分数": "score", "绩点": "gpa", "是否选修": "is_elective",
              "课程类别": "course_category", "修读类别": "retake_type",
              "重修重考": "retake_type"}


@dataclass
class ScoreRow:
    year: str
    semester: str
    student_no: str
    name: str
    class_name: str
    course_code: str
    course_name: str
    teacher: str
    credit: float | None
    score_raw: str
    score_num: float | None
    gpa: float | None
    is_elective: str = ""
    course_category: str = ""
    retake_type: str = ""
    sheet: str = ""
    row_index: int = 0


@dataclass
class ParsedScoreData:
    rows: list[ScoreRow] = field(default_factory=list)
    exceptions: list[dict] = field(default_factory=list)
    year_names: list[str] = field(default_factory=list)
    class_names: list[str] = field(default_factory=list)

    @property
    def student_count(self) -> int:
        return len({r.student_no for r in self.rows})

    @property
    def course_count(self) -> int:
        return len({r.course_code for r in self.rows})


def _iter_sheet_rows(filename: str, data: bytes):
    """统一产出 (sheet_name, row_values[])，值归一为 str/float/None。"""
    if filename.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=data)
        for sh in wb.sheets():
            yield sh.name, [[_cell_value(sh.cell_type(r, c), sh.cell_value(r, c))
                             for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        for ws in wb.worksheets:
            yield ws.title, [[_norm_cell(ws.cell(r, c).value)
                              for c in range(1, ws.max_column + 1)]
                             for r in range(1, ws.max_row + 1)]


def _cell_value(cell_type, value):
    if cell_type == 0:
        return None
    if cell_type == 2:
        return float(value)
    if cell_type == 3:
        return str(value)
    return str(value).strip() if value is not None else None


def _norm_cell(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return str(value).strip()


def _as_text(v) -> str:
    """学号/代码类值 → 干净文本（整数浮点去掉 .0）。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _norm_semester(v) -> str:
    text = _as_text(v)
    return SEMESTER_ALIASES.get(text, SEMESTER_ALIASES.get(text[:1], ""))


def normalize_class_name(name: str) -> str:
    return name.strip().removesuffix("班").strip()


def infer_grade_name(class_name: str) -> str | None:
    """班级名 → 年级名。规则：4 位数字取前两位（建筑类2401→24级、城规2101→21级）；
    仅 2 位数字直接使用（计科23班→23级）；无数字返回 None。"""
    nums = re.findall(r"\d+", normalize_class_name(class_name))
    four = [n for n in nums if len(n) >= 4]
    if four:
        return f"{max(four, key=len)[:2]}级"
    two = [n for n in nums if len(n) == 2]
    if two:
        return f"{two[0]}级"
    return None


def infer_enrollment_year(grade_name: str) -> int:
    return 2000 + int(grade_name.rstrip("级"))


def parse_score_workbook(filename: str, data: bytes, conversion: dict[str, float]) -> ParsedScoreData:
    parsed = ParsedScoreData()
    seen_header = False
    for sheet_name, rows in _iter_sheet_rows(filename, data):
        header_idx, col = None, {}
        for i, row in enumerate(rows):
            texts = {_as_text(v): j for j, v in enumerate(row) if v is not None}
            if "学号" in texts and ("课程代码" in texts or "课程号" in texts):
                header_idx = i
                col = {HEADER_MAP[t]: j for t, j in texts.items() if t in HEADER_MAP}
                break
        if header_idx is None:
            continue
        seen_header = True

        def get(row: list, en: str):
            j = col.get(en)
            return row[j] if j is not None and j < len(row) else None

        for ri in range(header_idx + 1, len(rows)):
            row = rows[ri]
            student_no = _as_text(get(row, "student_no"))
            course_code = _as_text(get(row, "course_code"))
            if not student_no and not course_code:
                continue  # 空行
            exc_base = {"sheet": sheet_name, "row": ri + 1}
            if not student_no:
                parsed.exceptions.append({**exc_base, "type": "缺学号", "detail": f"课程 {course_code}"})
                continue
            if not course_code:
                parsed.exceptions.append({**exc_base, "type": "缺课程代码", "detail": f"学号 {student_no}"})
                continue
            if col.get("year_term") is not None:
                yt = _as_text(get(row, "year_term"))
                m = re.match(r"^(\d{4})-(\d{4})学年.?([春夏秋])", yt)
                year_name = f"{m.group(1)}-{m.group(2)}" if m else yt
                semester = _norm_semester(m.group(3) if m else yt)
            else:
                year_name = _as_text(get(row, "year"))
                semester = _norm_semester(get(row, "semester"))
            if not semester:
                parsed.exceptions.append({**exc_base, "type": "未知学期",
                                          "detail": f"{student_no} 学期「{semester}」，按秋季处理"})
                semester = "秋季"

            credit = parse_number(get(row, "credit"))
            if credit is None:
                parsed.exceptions.append({**exc_base, "type": "缺学分",
                                          "detail": f"{student_no} {course_code}（不计入加权统计）"})
            gpa = parse_number(get(row, "gpa"))
            if gpa is None:
                parsed.exceptions.append({**exc_base, "type": "缺绩点", "detail": f"{student_no} {course_code}"})

            score_v = get(row, "score")
            score_raw = _as_text(score_v)
            score_num = parse_number(score_v)
            if score_num is None and score_raw:
                score_num = convert_level(score_raw, conversion)
                if score_num is None:
                    parsed.exceptions.append({**exc_base, "type": "未知等级",
                                              "detail": f"{student_no} {course_code} 分数「{score_raw}」（保留入库，不计入统计）"})

            parsed.rows.append(ScoreRow(
                year=year_name, semester=semester,
                student_no=student_no, name=_as_text(get(row, "name")) or "",
                class_name=normalize_class_name(_as_text(get(row, "class_name"))),
                course_code=course_code, course_name=_as_text(get(row, "course_name")) or "",
                teacher=_as_text(get(row, "teacher")) or "", credit=credit,
                score_raw=score_raw, score_num=score_num, gpa=gpa,
                is_elective=_as_text(get(row, "is_elective")) or "",
                course_category=_as_text(get(row, "course_category")) or "",
                retake_type=_as_text(get(row, "retake_type")) or "",
                sheet=sheet_name, row_index=ri + 1,
            ))
    if not seen_header:
        raise ValueError("未找到含「学号」「课程代码」的表头行")
    parsed.year_names = sorted({r.year for r in parsed.rows if r.year})
    parsed.class_names = sorted({r.class_name for r in parsed.rows if r.class_name})

    by_no: dict[str, set[str]] = defaultdict(set)
    for r in parsed.rows:
        if r.class_name:
            by_no[r.student_no].add(r.class_name)
    for no, classes in sorted(by_no.items()):
        if len(classes) > 1:
            parsed.exceptions.append({"sheet": "-", "row": "-", "type": "班级归属冲突",
                                      "detail": f"学号 {no} 在文件中出现多个班级：{'、'.join(sorted(classes))}"})
    return parsed


def find_conflicts_with_db(db: Session, parsed: ParsedScoreData) -> list[dict]:
    """文件姓名/班级 与系统不一致的冲突清单（人工裁决用）。"""
    nos = {r.student_no for r in parsed.rows}
    if not nos:
        return []
    students = {s.student_no: s for s in db.query(Student).filter(Student.student_no.in_(nos)).all()}
    file_info: dict[str, tuple[str, str]] = {}
    for r in parsed.rows:
        if r.class_name and (r.student_no not in file_info or r.name):
            file_info[r.student_no] = (r.name, r.class_name)
    conflicts = []
    for no, (name, class_name) in sorted(file_info.items()):
        s = students.get(no)
        if s is None:
            continue
        diffs = []
        if name and s.name != name:
            diffs.append(f"姓名 系统「{s.name}」/ 文件「{name}」")
        sys_class = s.klass.name if s.klass else ""
        if class_name and sys_class != class_name:
            diffs.append(f"班级 系统「{sys_class}」/ 文件「{class_name}」")
        if diffs:
            conflicts.append({"student_no": no, "name": s.name, "file_name": name,
                              "system_class": sys_class, "file_class": class_name,
                              "detail": "；".join(diffs)})
    return conflicts


def confirm_score_import(db: Session, parsed: ParsedScoreData, plan: dict,
                         user, filename: str) -> ImportBatch:
    """plan: {create_years:[{name}], create_grades:[{name,enrollment_year}],
              create_classes:[{name,grade_name,college_name}], conflicts:[{student_no,resolve}]}
    user: 操作人（辅导员新建的年级自动与其绑定，§4.1.3）。"""
    from ..models import College
    for y in plan.get("create_years", []):
        if not db.query(AcademicYear).filter_by(name=y["name"]).first():
            db.add(AcademicYear(name=y["name"]))
    for g in plan.get("create_grades", []):
        if not db.query(Grade).filter_by(name=g["name"]).first():
            grade = Grade(name=g["name"], enrollment_year=g["enrollment_year"])
            if user is not None and user.role == "counselor":
                user.grades.append(grade)
            db.add(grade)
    db.flush()
    for c in plan.get("create_classes", []):
        if db.query(ClassInfo).filter_by(name=c["name"]).first():
            continue
        grade = db.query(Grade).filter_by(name=c["grade_name"]).first()
        college = None
        if c.get("college_name"):
            college = db.query(College).filter_by(name=c["college_name"]).first()
            if college is None:
                college = College(name=c["college_name"])
                db.add(college)
                db.flush()
        db.add(ClassInfo(name=c["name"], grade_id=grade.id,
                         college_id=college.id if college else None))
    db.flush()

    resolve_map = {c["student_no"]: c.get("resolve", "system") for c in plan.get("conflicts", [])}
    year_ids = {y.name: y.id for y in db.query(AcademicYear).all()}
    class_ids = {c.name: c.id for c in db.query(ClassInfo).all()}
    students: dict[str, Student] = {s.student_no: s for s in db.query(Student).all()}

    batch = ImportBatch(kind="score", filename=filename,
                        operator_id=user.id if user is not None else None)
    db.add(batch)
    db.flush()

    # 同（学号+学年+学期+课程代码）文件内以后出现者为准
    dedup: dict[tuple, ScoreRow] = {}
    for r in parsed.rows:
        dedup[(r.student_no, r.year, r.semester, r.course_code)] = r

    stats = Counter()
    for (no, year, sem, code), r in dedup.items():
        year_id = year_ids.get(year)
        if year_id is None:
            stats["skipped_no_year"] += 1
            continue
        class_id = class_ids.get(r.class_name)
        student = students.get(no)
        if student is None:
            if not class_id:
                stats["skipped_no_class"] += 1
                continue
            student = Student(student_no=no, name=r.name or no, class_id=class_id)
            db.add(student)
            db.flush()
            students[no] = student
            stats["students_created"] += 1
        elif resolve_map.get(no) == "file":
            if r.class_name and class_id and student.class_id != class_id:
                student.class_id = class_id
                stats["students_moved"] += 1
            if r.name and student.name != r.name:
                student.name = r.name
        existing = db.query(ScoreRecord).filter_by(
            student_id=student.id, academic_year_id=year_id, semester=sem, course_code=code).first()
        fields = dict(course_name=r.course_name, teacher=r.teacher, credit=r.credit,
                      score_raw=r.score_raw, score_num=r.score_num, gpa=r.gpa,
                      is_elective=r.is_elective, course_category=r.course_category,
                      retake_type=r.retake_type, batch_id=batch.id)
        if existing:
            batch.snapshot.append({"model": "ScoreRecord", "id": existing.id, "old": {
                k: getattr(existing, k) for k in
                ["course_name", "teacher", "credit", "score_raw", "score_num", "gpa",
                 "is_elective", "course_category", "retake_type", "batch_id"]}})
            for k, v in fields.items():
                setattr(existing, k, v)
            stats["records_overwritten"] += 1
        else:
            db.add(ScoreRecord(student_id=student.id, academic_year_id=year_id,
                               semester=sem, course_code=code, **fields))
            stats["records_created"] += 1

    batch.stats = dict(stats)
    db.add(OperationLog(operator_id=user.id if user is not None else None, action="成绩导入",
                        detail=f"{filename}：新建{stats['records_created']} 覆盖{stats['records_overwritten']} "
                               f"学生新建{stats['students_created']} 异常{len(parsed.exceptions)}"))
    db.commit()
    db.refresh(batch)
    return batch

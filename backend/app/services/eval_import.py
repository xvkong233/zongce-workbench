"""综测明细导入：名册/保存/批量保存、Excel 导入（兼容新旧格式）、复制上一学年、批量填充基础分。"""

import io
import re
from dataclasses import dataclass, field

import openpyxl
from sqlalchemy.orm import Session

from ..models import (AcademicYear, ClassInfo, EvalRecord, ImportBatch,
                      OperationLog, Student)
from .calc import resolve_scheme
from .convert import is_detail_mismatch, sum_detail_terms
from .score_import import _as_text, _norm_cell, normalize_class_name

ITEM_HEADER_STOP = ("总分", "合计", "确认签字", "签名")


@dataclass
class EvalStudentRow:
    student_no: str
    name: str
    class_name_raw: str
    items: dict[str, dict] = field(default_factory=dict)  # name -> {detail, score, soft_sum}


@dataclass
class ParsedEvalData:
    title: str = ""
    sheet: str = ""
    students: list[EvalStudentRow] = field(default_factory=list)
    item_names: list[str] = field(default_factory=list)   # 文件表头解析出的项目（数字剥离）
    item_maxes: dict[str, float] = field(default_factory=dict)  # 表头携带的满分提示
    class_keys: list[str] = field(default_factory=list)   # 文件中出现的班级原文


def _squash_class_key(raw: str) -> str:
    """班级名变体归一：去空白/「班」后缀/「类」字 → 建筑2402。"""
    text = re.sub(r"\s+", "", str(raw or ""))
    text = normalize_class_name(text)
    return text.replace("类", "")


def match_class(db: Session, raw_names: list[str]) -> dict[str, dict]:
    """原文班级 → {class_id, name, ambiguous:[candidates]}。"""
    all_classes = db.query(ClassInfo).all()
    by_key: dict[str, list] = {}
    for c in all_classes:
        by_key.setdefault(_squash_class_key(c.name), []).append(c)
    result = {}
    for raw in raw_names:
        cands = by_key.get(_squash_class_key(raw), [])
        if len(cands) == 1:
            result[raw] = {"class_id": cands[0].id, "name": cands[0].name, "ambiguous": []}
        else:
            result[raw] = {"class_id": None, "name": raw,
                           "ambiguous": [{"id": c.id, "name": c.name} for c in cands]}
    return result


def parse_eval_workbook(data: bytes) -> ParsedEvalData:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    parsed = ParsedEvalData()
    for ws in wb.worksheets:
        # 合并单元格：左上角坐标(1-based) -> 跨列数
        merge_span: dict[tuple[int, int], int] = {}
        for rng in ws.merged_cells.ranges:
            merge_span[(rng.min_row, rng.min_col)] = rng.max_col - rng.min_col + 1
        grid = [[_norm_cell(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]

        def eff(row0: int, col0: int):
            """(0-based) 单元格有效值：处理纵向合并（值在左上角）。"""
            for dr in range(row0, -1, -1):
                span = merge_span.get((dr + 1, col0 + 1))
                if span:
                    return grid[dr][col0], span
            return grid[row0][col0], 1

        # 定位表头行：含「学号」与「姓名」
        header_idx = None
        for i, row in enumerate(grid):
            texts = [_as_text(v) for v in row]
            if "学号" in texts and "姓名" in texts:
                header_idx = i
                break
        if header_idx is None:
            continue
        header = grid[header_idx]
        no_col = next(j for j, v in enumerate(header) if _as_text(v) == "学号")
        name_col = next(j for j, v in enumerate(header) if _as_text(v) == "姓名")
        class_col = next((j for j, v in enumerate(header) if _as_text(v) == "班级"), None)

        # 项目列组：从学号列之后扫描，合并单元格占多列（如「思想品德25」合并 E2:F2 → 明细列 E、得分列 F）
        item_cols: list[tuple[str, int, int]] = []  # (名称原文, 明细列0基, 得分列0基)
        j = max(no_col, name_col) + 1
        while j < len(header):
            value, span = eff(header_idx, j)
            text = _as_text(value)
            if not text:
                j += 1
                continue
            if any(stop in text for stop in ITEM_HEADER_STOP):
                break
            if text in ("序号", "姓名", "学号", "班级"):
                j += span or 1
                continue
            item_cols.append((text, j, j + 1))  # 明细列在项目表头首列（E3=加减分项），得分列在其右
            j += max(span or 2, 2)  # 每个项目至少占「明细+得分」两列
        if not item_cols:
            continue

        parsed.sheet = ws.title
        title_vals = [str(v) for v in grid[0] if v is not None]
        if title_vals:
            parsed.title = "".join(title_vals)
        for orig, _, _ in item_cols:
            name = re.sub(r"\d+(\.\d+)?$", "", orig).strip()   # 思想品德25 → 思想品德
            m = re.search(r"(\d+(?:\.\d+)?)$", orig.strip())
            if name not in parsed.item_names:
                parsed.item_names.append(name)
                if m:
                    parsed.item_maxes[name] = float(m.group(1))
        if class_col is not None:
            for row in grid[header_idx + 1:]:
                v = row[class_col] if class_col < len(row) else None
                if v and _as_text(v) not in parsed.class_keys:
                    parsed.class_keys.append(_as_text(v))

        for ri in range(header_idx + 1, len(grid)):
            row = grid[ri]
            student_no = _as_text(row[no_col]) if no_col < len(row) else ""
            name = _as_text(row[name_col]) if name_col < len(row) else ""
            if not student_no:
                continue
            if not re.search(r"\d", student_no):
                continue
            srow = EvalStudentRow(student_no=student_no, name=name,
                                  class_name_raw=_as_text(row[class_col]) if class_col is not None and class_col < len(row) else "")
            for orig, dcol, scol in item_cols:
                item_name = re.sub(r"\d+(\.\d+)?$", "", orig).strip()
                detail = str(row[dcol] or "") if dcol < len(row) else ""
                score = _norm_cell(row[scol]) if scol < len(row) else None
                try:
                    score_f = round(float(score), 2) if score is not None else None
                except (TypeError, ValueError):
                    score_f = None
                soft = sum_detail_terms(detail) if detail else None
                srow.items[item_name] = {"detail": detail.strip(), "score": score_f, "soft_sum": soft}
            parsed.students.append(srow)
    if not parsed.students:
        raise ValueError("未在文件中找到含「学号」「姓名」表头的数据行")
    return parsed


def confirm_eval_import(db: Session, parsed: ParsedEvalData, academic_year_id: int,
                        grade_id: int, resolve: dict[str, str],
                        operator_id: int | None, filename: str) -> ImportBatch:
    """按年级导入（v1.3.3）：仅以学号匹配「所选年级内」的学生并覆盖其同年学年的旧记录；
    文件班级列仅作参考、无需映射。未匹配者（含年级外学号）记异常跳过。
    resolve: 学号→'skip' 强制跳过。"""
    from ..models import Grade
    grade = db.get(Grade, grade_id)
    if grade is None:
        from fastapi import HTTPException
        raise HTTPException(400, {"message": "目标年级不存在"})
    batch = ImportBatch(kind="eval", filename=filename, academic_year_id=academic_year_id,
                        operator_id=operator_id)
    db.add(batch)
    db.flush()
    scheme = resolve_scheme(db, academic_year_id, grade_id)
    max_by_name = {i["name"]: i.get("max_score") for i in scheme.items}
    stats = {"students_updated": 0, "records_created": 0, "records_overwritten": 0,
             "unmatched": 0, "soft_mismatch": 0}
    unmatched_list = []
    nos = [s.student_no for s in parsed.students]
    students = {s.student_no: s for s in
                db.query(Student).join(ClassInfo).filter(
                    ClassInfo.grade_id == grade_id,
                    Student.student_no.in_(nos)).all()} if nos else {}
    for srow in parsed.students:
        student = students.get(srow.student_no)
        if student is None or resolve.get(srow.student_no) == "skip":
            stats["unmatched"] += 1
            unmatched_list.append({"student_no": srow.student_no, "name": srow.name,
                                   "class": srow.class_name_raw})
            continue
        for item_name, data in srow.items.items():
            score = data.get("score")
            if score is None:
                score = data.get("soft_sum") or 0.0
            if is_detail_mismatch(data.get("soft_sum"), score, max_by_name.get(item_name)):
                stats["soft_mismatch"] += 1
            existing = db.query(EvalRecord).filter_by(
                student_id=student.id, academic_year_id=academic_year_id, item_name=item_name).first()
            if existing:
                batch.snapshot.append({"model": "EvalRecord", "id": existing.id, "old": {
                    "detail_text": existing.detail_text, "score": existing.score,
                    "batch_id": existing.batch_id}})
                existing.detail_text = data["detail"]
                existing.score = score
                existing.batch_id = batch.id
                stats["records_overwritten"] += 1
            else:
                db.add(EvalRecord(student_id=student.id, academic_year_id=academic_year_id,
                                  item_name=item_name, detail_text=data["detail"],
                                  score=score, batch_id=batch.id))
                stats["records_created"] += 1
        stats["students_updated"] += 1
    batch.stats = {**stats, "unmatched_list": unmatched_list}
    db.add(OperationLog(operator_id=operator_id, action="综测导入",
                        detail=f"{filename}（{grade.name}）：{stats['students_updated']}人 "
                               f"新建{stats['records_created']} 覆盖{stats['records_overwritten']} "
                               f"未匹配{stats['unmatched']} 明细不符{stats['soft_mismatch']}"))
    db.commit()
    db.refresh(batch)
    return batch

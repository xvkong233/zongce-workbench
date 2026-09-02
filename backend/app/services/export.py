"""output.xlsx 风格导出：完整工作簿（每年级 成绩+绩点 两 sheet）与综测简表。

布局对齐手工样本 output.xlsx：
- A/B 列表头区留白；C 列 1..3(4) 行为「课程代码/课程名称/(授课教师)/学分」元标签；
- 课程列自 D 列起；表头末行为「班级|学号|姓名」标签行（A/B/C 列）；
- 数据行按班级分组、学号升序，稀疏矩阵；尾部统计列含公式与双排名。
"""
from __future__ import annotations

from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models import AcademicYear, Student
from .calc import StudentResult, compute_grade, resolve_scheme

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
RANK_FILL = PatternFill("solid", fgColor="FFF2CC")
MISSING_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FIRST_COURSE_COL = 4  # D 列起为课程（A/B 留白，C 列元标签）


def _majority(values: list) -> object:
    """出现次数最多者；平票取字典序首个保证稳定。"""
    counts = Counter(values)
    return max(counts.items(), key=lambda kv: (kv[1], str(kv[0])))[0]


def _course_columns(records: list) -> list[dict]:
    """课程列定义：秋季在前、春季在后，同学期按代码升序。"""
    by_code: dict[str, list] = {}
    for r in records:
        by_code.setdefault(r.course_code, []).append(r)
    cols = []
    for code, group in by_code.items():
        sem = _majority([r.semester for r in group]) or "秋季"
        cols.append({"code": code, "name": _majority([r.course_name for r in group]),
                     "teacher": _majority([r.teacher for r in group]) or "",
                     "credit": _majority([r.credit for r in group]),
                     "semester": sem})
    cols.sort(key=lambda c: (0 if c["semester"] == "秋季" else 1, c["code"]))
    return cols


def _kept_records(db, students, year_id, rule) -> list:
    from ..models import ScoreRecord
    from .calc import collapse_retakes
    kept = []
    for s in students:
        recs = db.query(ScoreRecord).filter_by(student_id=s.id, academic_year_id=year_id).all()
        kept.extend(collapse_retakes(recs, rule))
    return kept


def _write_eval_cells(ws, row, start_col, res: StudentResult, item_names: list[str]):
    """尾部统计列。未录综测者五项与综合测评显示 0 并标记填充色。"""
    col = start_col

    def put(value, fill=None, numfmt=None):
        nonlocal col
        cell = ws.cell(row=row, column=col, value=value)
        if fill:
            cell.fill = fill
        if numfmt:
            cell.number_format = numfmt
        col += 1
        return cell

    put(res.weighted_total, numfmt="0.00")
    put(res.total_credit, numfmt="0.00")
    if res.weighted_avg is not None:
        wt, tc = get_column_letter(start_col), get_column_letter(start_col + 1)
        put(f"={wt}{row}/{tc}{row}", numfmt="0.00")
    else:
        put(None)
    put(None)  # 德育留空
    for name in item_names:
        entered = res.items_entered.get(name)
        put(res.item_scores.get(name, 0), fill=None if entered else MISSING_FILL, numfmt="0.00")
    put(res.eval_total if res.eval_entered else 0,
        fill=None if res.eval_entered else MISSING_FILL, numfmt="0.00")
    if res.final_score is not None:
        put(res.final_score, numfmt="0.00")
    else:
        put(None)
    put(res.academic_rank, fill=RANK_FILL)
    put(res.eval_rank, fill=RANK_FILL)
    put(None)  # 确认签字留空


def _style_sheet(ws, header_rows: int, last_col: int, data_last_row: int):
    for r in range(1, header_rows + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.alignment = CENTER
            cell.border = BORDER
            cell.fill = HEADER_FILL
    for r in range(header_rows + 1, data_last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c != 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws.cell(row=header_rows + 1, column=FIRST_COURSE_COL)


def build_grade_sheet(ws, grade_name: str, students: list[Student], kind: str,
                      scheme, results: dict[int, StudentResult], kept: list | None = None):
    """kind: score|gpa。students 需已按（班级，学号）排序；kept 为收敛后的成绩记录。"""
    if kept is None:  # 兼容直接调用（真实流程由 build_workbook 传入）
        kept = []
    courses = _course_columns(kept)
    item_names = [i["name"] for i in scheme.items]

    meta_rows = 4 if kind == "gpa" else 3
    label_row = meta_rows + 1
    meta_labels = (["课程代码", "课程名称", "学分"] if kind == "score"
                   else ["课程代码", "课程名称", "授课教师", "学分"])
    for r, lab in enumerate(meta_labels, start=1):
        ws.cell(row=r, column=3, value=lab)
    ws.cell(row=label_row, column=1, value="班级")
    ws.cell(row=label_row, column=2, value="学号")
    ws.cell(row=label_row, column=3, value="姓名")
    ws.merge_cells(start_row=1, start_column=1, end_row=meta_rows, end_column=2)

    for ci, course in enumerate(courses):
        col = FIRST_COURSE_COL + ci
        ws.cell(row=1, column=col, value=course["code"])
        ws.cell(row=2, column=col, value=course["name"])
        if kind == "gpa":
            ws.cell(row=3, column=col, value=course["teacher"])
            ws.cell(row=4, column=col, value=course["credit"])
        else:
            ws.cell(row=3, column=col, value=course["credit"])

    tail_start = FIRST_COURSE_COL + len(courses)
    tail_labels = ["加权总分", "总学分", "加权平均值", "德育", *item_names,
                   "综合素质测评", "综合测评成绩", "智育排名", "综测排名", "确认签字"]
    for i, lab in enumerate(tail_labels):
        ws.cell(row=label_row, column=tail_start + i, value=lab)

    rec_index = {(r.student_id, r.course_code): r for r in kept}
    row = label_row + 1
    for s in students:
        ws.cell(row=row, column=1, value=s.klass.name if s.klass else "")
        ws.cell(row=row, column=2, value=int(s.student_no) if s.student_no.isdigit() else s.student_no)
        ws.cell(row=row, column=3, value=s.name)
        res = results[s.id]
        for ci, course in enumerate(courses):
            rec = rec_index.get((s.id, course["code"]))
            if rec is None:
                continue
            value = rec.score_num if kind == "score" else rec.gpa
            if value is not None:
                ws.cell(row=row, column=FIRST_COURSE_COL + ci, value=value)
        _write_eval_cells(ws, row, tail_start, res, item_names)
        row += 1

    last_col = tail_start + len(tail_labels)
    _style_sheet(ws, label_row, last_col, row - 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 10
    for ci in range(len(courses)):
        ws.column_dimensions[get_column_letter(FIRST_COURSE_COL + ci)].width = 7
    for i in range(len(tail_labels)):
        ws.column_dimensions[get_column_letter(tail_start + i)].width = 9
    ws.title = f"{grade_name}（{'成绩' if kind == 'score' else '绩点'}）"


def build_brief_sheet(ws, grade_name: str, students, scheme, results):
    item_names = [i["name"] for i in scheme.items]
    labels = ["班级", "学号", "姓名", "学业加权平均", *item_names,
              "综合素质测评", "综合测评成绩", "智育排名", "综测排名", "确认签字"]
    for c, lab in enumerate(labels, start=1):
        ws.cell(row=1, column=c, value=lab)
    row = 2
    for s in students:
        res = results[s.id]
        vals = [s.klass.name if s.klass else "",
                int(s.student_no) if s.student_no.isdigit() else s.student_no, s.name,
                res.weighted_avg if res.weighted_avg is not None else "—"]
        for name in item_names:
            vals.append(res.item_scores.get(name) if res.eval_entered else "未录入")
        vals += [res.eval_total if res.eval_entered else "未录入",
                 res.final_score if res.final_score is not None else "—",
                 res.academic_rank if res.academic_rank is not None else "—",
                 res.eval_rank if res.eval_rank is not None else "—", None]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        row += 1
    _style_sheet(ws, 1, len(labels), row - 1)
    ws.title = f"{grade_name}综测简表"


def build_workbook(db, year: AcademicYear, grades, class_ids: list[int] | None = None,
                   brief: bool = False) -> Workbook:
    """grades: [(grade, students)]，students 已按（班级名，学号）排序。"""
    wb = Workbook()
    wb.remove(wb.active)
    for grade, students in grades:
        scheme = resolve_scheme(db, year.id, grade.id)
        results = {r.student_id: r for r in compute_grade(db, students, year.id, scheme)}
        kept = _kept_records(db, students, year.id, scheme.retake_rule or "latest")
        if brief:
            ws = wb.create_sheet()
            build_brief_sheet(ws, grade.name, students, scheme, results)
        else:
            ws1 = wb.create_sheet()
            build_grade_sheet(ws1, grade.name, students, "score", scheme, results, kept)
            ws2 = wb.create_sheet()
            build_grade_sheet(ws2, grade.name, students, "gpa", scheme, results, kept)
    return wb


def sort_students(students: list[Student]) -> list[Student]:
    return sorted(students, key=lambda s: (s.klass.name if s.klass else "", s.student_no))
